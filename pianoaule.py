import requests
import openpyxl
import csv
import re
import json
import os
from datetime import datetime
import logging
import argparse
import excel2img
###############
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
#
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw
# Argument Parser
WORKING_DIR = f"{os.path.abspath(os.getcwd())}/output"
os.makedirs(WORKING_DIR, exist_ok=True)
os.chdir(WORKING_DIR)
FILENAME = "PianoAuleByVI"

parser = argparse.ArgumentParser(epilog="Piano Aule by Vivere Ingegneria", description="\
Questo script permette di creare automaticamente una serie di file in diversi\
formati contenenti il piano aule, il file finale contenente l'immagine definitiva si trovera' nella cartella Screen")

parser.add_argument("--week-offset", "-w", type=int, default=0, metavar="01", 
                    help="Per indicare quale settimana analizzare rispetto a quella corrente")
parser.add_argument("--multi-file", "-m", default=False, action="store_true", 
                    help="Crea un file csv per ogni giorno della settimana.")
parser.add_argument("--cleanup", "-c", choices=["non_png", "csv", "xlsx", "png", "all"], action="append",
                    help="Se vuoi cancellare i file creati automaticamente.", )
parser.add_argument("--verbose", "-v", action="store_true", default=False, help="Se vuoi vedere dettagli sullo stato dello script")


args = parser.parse_args()
SINGLE_FILE = not args.multi_file
WEEK_OFFSET = args.week_offset
CLEANUP = args.cleanup
logging.basicConfig(level=logging.ERROR)
lvl = logging.DEBUG if args.verbose else logging.INFO
logger = logging.getLogger(name="Piano Aule by VI")
# Logging
logger.setLevel(lvl)
#

# Session
session = requests.Session()
#

AULE_OID = {
    "A210": "798",
    "A220": "192",
    "A310": "151",
    "A320": "152",
    "B110": "195",
    "B120": "196",
    "B210": "197",
    "B310": "198",
    "F010": "304",
    "F100": "243",
    "F110": "244",
    "F120": "307",
    "F130": "322",
    "F140": "313",
    "F150": "314",
    "F160": "323",
    "F170": "324",
    "F180": "315",
    "F190": "325",
    "F210": "308",
    "F220": "321",
    "F230": "309",
    "F240": "310",
    "F310": "311",
    "F320": "312",
    "G210": "186",
    "G220": "187",
    "L010": "227",
    "M010": "327",
    "M020": "199",
    "N020": "326",
    "N030": "200",
    "N040": "201",
    "O010": "269",
    "O011": "281",
    "O012": "282",
    "T110": "212",
    "T120": "213",
    "T210": "767",
    "T220": "215",
    "T230": "221",
    "U110": "216",
    "U120": "217",
    "U140": "218",
    "U150": "219",
    "U160": "220",
    "U170": "241",
}

GIORNI_DELLA_SETTIMANA = ["LUN", "MAR", "MER", "GIO", "VEN", "SAB"]
HEADERS = [
    'AULA', 'LUN 08:00', 'LUN 09:00', 'LUN 10:00', 'LUN 11:00', 'LUN 12:00', 'LUN 13:00', 'LUN 14:00',
    'LUN 15:00', 'LUN 16:00', 'LUN 17:00', 'LUN 18:00', 'MARTEDÌ', 'MAR 08:00', 'MAR 09:00', 'MAR 10:00', 'MAR 11:00',
    'MAR 12:00', 'MAR 13:00', 'MAR 14:00', 'MAR 15:00', 'MAR 16:00', 'MAR 17:00', 'MAR 18:00', 'MERCOLEDÌ', 'MER 08:00',
    'MER 09:00', 'MER 10:00', 'MER 11:00', 'MER 12:00', 'MER 13:00', 'MER 14:00', 'MER 15:00', 'MER 16:00',
    'MER 17:00', 'MER 18:00', 'GIOVEDÌ', 'GIO 08:00', 'GIO 09:00', 'GIO 10:00', 'GIO 11:00', 'GIO 12:00', 'GIO 13:00',
    'GIO 14:00', 'GIO 15:00', 'GIO 16:00', 'GIO 17:00', 'GIO 18:00', 'VENERDÌ', 'VEN 08:00', 'VEN 09:00', 'VEN 10:00',
    'VEN 11:00', 'VEN 12:00', 'VEN 13:00', 'VEN 14:00', 'VEN 15:00', 'VEN 16:00', 'VEN 17:00', 'VEN 18:00']


def get_calendar_from_oid(oid: str, **kwargs):
    """
    Questa funzione fa la richiesta ad unipa per ottenere il calendario di una determinata aula
    :param oid: oid dell'aula
    :return: una stringa che sarebbe la variabile events del sito unipa, la stringa è già pulita
    """
    url = "https://offweb.unipa.it/offweb/public/aula/calendar.seam?oidAula={}".format(oid)
    response = session.get(url)

    if not response.ok:
        raise ConnectionError("Errore nella risposta del server")

    logger.debug("Risposta: {}".format(response.ok))
    c = re.findall("var events = (.*?);", str(response.content))[0]
    return clean_calendar_response(c)


def clean_calendar_response(response: str):
    """
    Pulisce la risposta da offweb con una stringa parsabile dalla libreria json. un bordello che proverò a spiegare
    :param response: la stringa sporchissima che viene direttamente da unipa
    :return: stringa pulita, ancora da convertire in dict
    """
    # levo tutti i caratteri che danno fastidio

    logger.debug("Pulendo e inserendo i doppi apici")
    response = response.replace("\\t", '').replace("\\n", '').replace("\\", '').replace("'", '"')
    # queste parole devo racchiuderle in doppi apici
    keywords = ["className:", 'start:', 'end:', 'allDay:']
    for word in keywords:
        # racchiudo queste parole in doppi apici cosi da evitare problemi
        response = response.replace(word, '"' + word[:-1] + '":')
    # questo è per cancellare una informazione che non ci serve e che crea problemi
    to_delete = re.findall('title: .*?",', response)
    for item in to_delete:
        response = response.replace(item, "")
    logger.debug("Rimosso i title")

    # qua racchiudiamo le variabili contenenti il tempo in formato UNIX
    response = response.replace("new Date", '"new Date').replace("0),", '0)",')
    logger.debug("Racchiuso le date in unix time")
    # questo serve per rimuovere l'ultima virgola che fa problemi con la libreria json quando si converte
    for i in range(100):
        if  i<len(response) and response[-i] == ",":
            response = response[:-i] + response[-i + 1:]
            break
    logger.debug("Finito di pulire la stringa")
    return response


def parse_dates(calendar_list: list, **kwargs: dict):
    """
    Rimpiazza la data in unix time con un dict contenente weekday e hour
    weekday rappresenta il giorno della settimana (0 - Lun, 6 - Dom)
    hour rappresenta l' ora della giornata
    :param calendar_list:
    :return:
    """
    logger.debug("Cominciando a convertire le date da unix time...")
    week_offset = kwargs.get("week_offset") if kwargs.get("week_offset") is not None else WEEK_OFFSET
    for lecture in calendar_list:
        unix_start = lecture.get("start").split("(")[1].split(")")[0][:10]
        unix_end = lecture.get("end").split("(")[1].split(")")[0][:10]
        start_date = datetime.fromtimestamp(int(unix_start))
        end_date = datetime.fromtimestamp(int(unix_end))
        # logger.debug("Start UNIX: {} | Date: {}".format(unix_start, start_date.strftime("%c")))
        # logger.debug("End   UNIX: {} | Date: {}".format(unix_end, end_date.strftime("%c")))

        # questa parte qui serve per filtrare il calendario e effettivamente utilizzare solo le lezioni della stessa
        # settimana in cui viene runnato lo script
        today = datetime.now()
        logger.debug("Sto confrontando le date, risultato {} != {} and {} != {}".format(start_date.isocalendar()[1],
                                                                                         today.isocalendar()[1],
                                                                                         start_date.year,
                                                                                         today.year))
        if start_date.isocalendar()[1] != today.isocalendar()[1] + week_offset or start_date.year != today.year:
            # logger.debug("Quest'ultimo non era della settimana e dell'anno che stiamo analizzando, quindi ignorato")
            logger.debug("Scartato: " + start_date.strftime("%d-%m-%Y"))
            continue
        logger.debug("Data nella settimana corrente: " + start_date.strftime("%d-%m-%Y"))
        # questo if ignora tutti i giorni che non sono tra lun e ven
        if start_date.weekday() > 4:
            logger.debug("Quest'ultimo non era compreso tra Lunedì e Venerdì")
            continue

        else:
            lecture["start"] = {
                'weekday': start_date.weekday(),
                'hour': start_date.hour,
            }
            lecture["end"] = {
                'weekday': end_date.weekday(),
                'hour': end_date.hour,
            }
            logger.debug("Lezione: {}".format(lecture))
    # Rimuovo tutte le date ignorate
    calendar_list = [lecture for lecture in calendar_list if isinstance(lecture["start"], dict)]
    logger.debug("Parsing delle date finito.")
    return calendar_list


def ready_to_use_dict_calendar(oid, **kwargs):
    """
    Questa funzione prende in input l'oid e restituisce il dizionario con le lezioni della settimana corrente
    Non dovrebbero esserci problemi fin qui
    :param oid:
    :return:
    """
    clean_c = get_calendar_from_oid(oid, **kwargs)
    cal = json.loads(clean_c)
    return parse_dates(cal, **kwargs)


def get_conditional_formatting_rules(free_color: str = "6cae85", busy_color: str = "c6143a"):
    # definisco i colori per la formattazione condizionale
    red_fill = PatternFill(bgColor=busy_color)
    red_font = Font(color=busy_color)
    green_fill = PatternFill(bgColor=free_color)
    green_font = Font(color=free_color)
    white_font = Font(color="ffffff")

    # formattazioni condizionali in base alle etichette che abbiamo inserito in fase di scrittura
    # fill indica il background, font il colore del font
    # per oscurare il testo utilizziamo stesso colore per font e background
    rule_occ = Rule(type="containsText", operator="containsText", text="OCC",
                    dxf=DifferentialStyle(fill=red_fill, font=red_font))
    rule_occ.formula = ['NOT(ISERROR(SEARCH("OCC",A1)))']

    rule_free = Rule(type="containsText", operator="containsText", text="FREE",
                     dxf=DifferentialStyle(fill=green_fill, font=green_font))
    rule_free.formula = ['NOT(ISERROR(SEARCH("FREE",A1)))']

    rule_day = Rule(type="containsText", operator="containsText", text="DAY", dxf=DifferentialStyle(font=white_font))
    rule_day.formula = ['NOT(ISERROR(SEARCH("DAY",A1)))']

    return [rule_occ, rule_free, rule_day]


def get_calendars_from_unipa(aule = AULE_OID, **kwargs):
    calendari = {}
    total = len(aule.items())
    index = 0
    for nome_aula, oid_aula in aule.items():
        index += 1
        # per ogni aula
        # nome_aula equivale al nome comune (ES. F130)
        # oid_aula equivale al suo identificativo univoco (ES. 322)
        if not calendari.get(nome_aula):
            logger.info(f"Scaricando le lezioni nell'aula: {nome_aula} ({index}/{total})")
            calendari[nome_aula] = ready_to_use_dict_calendar(oid_aula, **kwargs)
    return calendari

def create_final_csv(filename, calendari: dict, **kwargs: dict):
    """
    Questa funzione dovrebbe essere quella che fetcha i dati dal sito e crea il file csv con tutte le informazioni.
    :param filename:
    :param aule:
    :return filenames: filenames used
    """
    # inizializzo gli oggetti che mi permettono di scrivere il csv
    single_file = kwargs.get("single_file") if kwargs.get("single_file") is not None else SINGLE_FILE
    filenames = [filename]
    if not single_file:
        filenames = [str(filename[:-3] + d + ".csv") for d in GIORNI_DELLA_SETTIMANA[:5]]

    for giorno, filename in enumerate(filenames):
        if single_file:
            current_headers = HEADERS
        else:
            current_headers = ["AULA"]
            current_headers.extend(d for d in HEADERS if d.startswith(GIORNI_DELLA_SETTIMANA[giorno])
                                   and d.endswith("00"))
        logger.info("Scrivendo il file csv...")
        c = open(filename, "w") 
        writer = csv.DictWriter(c, fieldnames=current_headers, lineterminator="\n")
        writer.writeheader()            
        for nome_aula, calendario in calendari.items():
            # row è quella variabile che alla fine andrò a scrivere sul file csv, corrisponde effettivamente ad una riga
            # del file csv
            # row è un dizionario, ogni valore verrà posizionato nella colonna che ha il nome della sua chiave
            # in questo caso nome_aula verrà posizionato nella colonna che ha come nome "AULA"
            row = {"AULA": nome_aula}
            for lezione in calendario:
                # per ogni lezione della aula che stiamo prendendo in considerazione ( cella di una determinata riga )
                # mi salvo gli orari della lezione
                start_wd = lezione["start"]["weekday"]
                start_h = lezione["start"]["hour"]
                end_h = lezione["end"]["hour"]
                if start_wd != giorno and not single_file:
                    # se non combacia con il giorno del file (in caso sia in MULTI_FILE mode)
                    continue
                day = str(GIORNI_DELLA_SETTIMANA[start_wd])
                logger.debug(f"WD: {start_wd} | DAY: {day} | START: {start_h} | END: {end_h}")
                end_h = 19 if end_h > 19 else end_h
                for i in range(start_h, end_h):
                    # aggiungo la lezione alla riga csv per tutta la durata di quest'ultima
                    # come chiave di questa entry metto il nome del giorno + l'ora della lezione
                    # (se c'è bisogno gli metto uno 0 prima) + :00
                    # Esempio: "LUN" + " " + "09" + ":00" => "LUN 09:00" che combacia con il nome di una colonna già
                    # esistente
                    row.update({day + " " + str(i).rjust(2, "0") + ":00": "OCC"})

            row.update({day: nome_aula for day in current_headers if day.endswith("Ì")})
            for h in current_headers:
                if not row.get(h):
                    row[h] = "FREE"
            logger.debug("Aggiunta la seguente riga: {}".format(row))
            writer.writerow(row)
        c.close()
    return [f[:-4] for f in filenames]


def csv_to_xlsx(csv_basename: str, xlsx_basename: str = None):
    logger.info("Convertendo il file CSV in excel...")

    wb = openpyxl.Workbook()
    ws = wb.active

    # elimina i bordi delle celle
    # ws.sheet_view.showGridLines = False

    # crea il csv e separa le celle tramite virgole
    csv_filename = csv_basename if csv_basename.endswith(".csv") else csv_basename + ".csv"
    with open(csv_filename) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    # scorriamo le colonne per aumentarne la dimensione (85px = 12cm)
    # scorriamo anche le righe in modo tale da fare l'allineamento al centro
    for i in range(1, 70):
        ws.column_dimensions[get_column_letter(i)].width = 12
        for j in range(1, 49):
            ws.cell(row=j, column=i).alignment = Alignment(horizontal="center")
            ws.cell(row=j, column=i).font = Font(name="Segoe UI", size=11)

    # Applico le regole di formattazione
    rules = get_conditional_formatting_rules()
    cell_range = "A1:BI48"
    for rule in rules:
        ws.conditional_formatting.add(cell_range, rule)

    filename_output = (xlsx_basename or csv_basename) + '.xlsx' 
    wb.save(filename_output)
    return xlsx_basename or csv_basename


def xlsx_to_png(xlsx_basename: str, png_basename: str = None, cell_range: str = "A1:BH48"):
    # conversione da xlsx a bmp tramite range d'azione
    logger.info(f"Convertendo il file {png_basename} excel in png...")
    bmp_filename = f"{xlsx_basename}.bmp"
    excel2img.export_img(xlsx_basename + '.xlsx', bmp_filename, "", f"Sheet!{cell_range}")
    piano = Image.open(bmp_filename)
    piano.save(f"{png_basename or xlsx_basename}.png")
    

def cleanup(**kwargs: dict):
    cleanup = kwargs.get("cleanup") if kwargs.get("cleanup") else CLEANUP
    if "non_png" in cleanup:
        to_clean = ["csv", "xlsx", "bmp"]
    elif "all" in cleanup: 
        to_clean = ["csv", "xlsx", "bmp", "png"]
    else: 
        to_clean = cleanup
    to_clean.append("bmp") # eliminarlo sempre, è inutile
    files = [f for f in os.listdir() if f.endswith(tuple(to_clean))]
    for f in files:
        os.remove(f)


def generate_pngs(**kwargs: dict):
    single_file = kwargs.get("single_file") if kwargs.get("single_file") is not None else SINGLE_FILE

    calendari = get_calendars_from_unipa(aule=AULE_OID, **kwargs)
    csvs = create_final_csv(filename=FILENAME + ".csv", calendari=calendari, **kwargs)
    for csv in csvs:
        xlsx_basename = csv_to_xlsx(csv_basename=csv)
        if not single_file:
            cell_range = "A1:L48"
        else: cell_range = "A1:BH48"
        xlsx_to_png(xlsx_basename, cell_range=cell_range)
    logger.info("Pulendo i file inutili...")
    cleanup(**kwargs)


if __name__ == '__main__':
    logger.info("Starting...")
    generate_pngs()